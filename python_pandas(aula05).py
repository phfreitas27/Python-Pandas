import matplotlib.pyplot as plt

plt.hist(combustiveis_df['Valor de Venda'])

plt.title("Preço dos combustíveis - Nov/2021")

plt.xlabel("Preço (em reais)")
plt.ylabel("Quantidade de coletas")

plt.axvline(combustiveis_df['Valor de Venda'].mean(), color='red', linestyle = 'dashed', linewidth = '1')

plt.show()

c_mean= combustiveis_df['Valor de Venda'].groupby(by=combustiveis_df['Produto']).mean()
display(c_mean)

import seaborn as sns

plt.figure(figsize=(10,10))

c_mean_grafico= c_mean.plot(kind="barh",
   title = "Preço Médio", 
   color = "red",
   alpha = 0.2)

c_mean_grafico.set_ylabel("Tipo de Combustivel")
c_mean_grafico.set_xlabel("Preço reais/litro")


plt.grid()

sns.despine(top = True, right= True, left = True, bottom = True,)

plt.show()

#c_mean.reset_index(inplace = True)

display(c_mean)

excel = "por_litro.xlsx"
c_mean.to_excel(excel, "Sumario")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment

wb = load_workbook(excel)

ws = wb['Sumario']

cinza = PatternFill("solid", fgColor="808080")
coords = ['A1', 'B1']
for coord in coords:
 ws[coord].fill = cinza

MAX_ROW = ws.max_row
num_linha = 2
while(num_linha <= MAX_ROW):
  coord = 'B' + str(num_linha)
  if ws[coord].value >= 6.5:
    ws[coord].font = Font(bold = True, color= "FF0000")
  num_linha = num_linha + 1
  
wb.save(excel)